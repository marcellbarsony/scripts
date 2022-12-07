#!/usr/bin/env python3

from openpyxl import Workbook, load_workbook

wb = Workbook()
#wb = load_workbook('test.xlsx')
ws = wb.active

#ws2 = wb.create_sheet("Sheet2")
#wb = wb.active
#print(wb)

wb.save(filename="test.xlsx")
